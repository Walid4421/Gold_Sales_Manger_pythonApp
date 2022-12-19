
import xlsxwriter # this is used for letting my program write in xlsx files
from tkinter import * # this is for making the UI
from openpyxl import Workbook, load_workbook
import datetime# for date and time
import os


from functions import *

date = datetime.datetime.now()
root = Tk()
root.title("DATA ENTRY")
root.geometry("400x500")
root.configure(bg="#000000")


file = "Sales.xlsx"

if os.path.isfile(file) == FALSE: # this is to check whether there is a sales xlsx file for input and if not it would create a file named Sales.xlsx
    workbook = xlsxwriter.Workbook(file)
    worksheet = workbook.add_worksheet("Month " + str(date.date().month) + " - " + str(date.date().year))# creates the first work sheet with the month and year as its name

    worksheet.write('A1', 'DATE')
    worksheet.write('B1', 'BUYER')
    worksheet.write('C1', 'PRODUCT NAME')
    worksheet.write('D1', 'PRICE(TRY)')
    worksheet.write('E1', 'WEIGHT(Gram)')
    worksheet.write('F1', 'Gold Price(gram-USD)')
    worksheet.write('G1', 'Gold Price(gram-TRY)')
    worksheet.write('H1', 'Dollar to Turkish Lira')

#set_column function enables on to change the size of the boxes.

    worksheet.set_column(0, 0, 13)
    worksheet.set_column(1, 1, 20)
    worksheet.set_column(2, 2, 30)
    worksheet.set_column(3, 3, 13)
    worksheet.set_column(4, 4, 15)
    worksheet.set_column(5, 5, 20)
    worksheet.set_column(6, 6, 20)
    worksheet.set_column(7, 7, 17)

    workbook.close()

#loads the file

wb = load_workbook(file)
ws = wb.active

#checks for a file that contains the last date the program was used
if os.path.isfile('Date.txt') == FALSE: #if there wasnt any flie like that then it creates one
    f = open('Date.txt', 'w')
    f.write(str(date.date()))
    f.close()
else:
    f = open('Date.txt', 'r')
    f_date = f.read(11)
    f.close()
    if int(f_date[5:7]) < date.date().month or int(f_date[:4]) < date.date().year: # this statement checks whether the month has changed or the year has changed
        ws.append([' '])
        ws.append(['TOTAL SOLD ', '=SUM(D:D)'])
        ws.append(['TOTAL WEIGHT ', '=SUM(E:E)'])
        f = open('Date.txt', 'w')
        wb.create_sheet("Month " + str(date.date().month) + " - " + str(date.date().year), 0)
        wb.save(file)
        ws = wb.active
        ws.append(['DATE', 'BUYER', 'PRODUCT NAME', 'PRICE(TRY)', 'WEIGHT(Gram)', 'Gold Price(gram-USD)',
                   'Gold Price(gram-TRY)', 'Dollar to Turkish Lira'])
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 21
        ws.column_dimensions["C"].width = 31
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 21
        ws.column_dimensions["G"].width = 21
        ws.column_dimensions["H"].width = 21
        f.write(str(date.date()))
        f.close()


def clicked():# this function is to take all the input and add it to the xlsx file

    name_e = e.get()
    product = e1.get()
    # this is used to try to assign the data to weight and if it was unsuccessful it would print an error to the user
    try:
        float(e2.get())
        weight = e2.get()
    except ValueError:
        err = Label(root, text=" Error(weight) ", bg="#000000", fg="White")
        err.grid(column=3)

    # this is used to try to assign the data to price and if it was unsuccessful it would print an error to the user
    try:
        float(e3.get())
        price_e = e3.get()
    except ValueError:
        err = Label(root, text=" Error(price) ", bg="#000000", fg="White")
        err.grid(column=3)

    ws.append([date.date(), name_e, product, float(price_e), float(weight),
               float(html[index_val_update: index_val_update_end]) / 31.1034768,
               (float(html[index_val_update: index_val_update_end])/31.1034768) * nae, nae])
    wb.save(file)
    done = Label(root, text=" DONE ", bg="#000000", fg="White")
    done.grid(column=3)
    e.delete(0, END)
    e1.delete(0, END)
    e2.delete(0, END)
    e3.delete(0, END)


space1 = Label(root, text="           ", bg="#000000")
space2 = Label(root, text="           ", bg="#000000")
space = Label(root, text="   ", bg="#000000")
space.grid(row=2)
space2.grid(row=9, column=50)
space1.grid(row=0)


# creating the name
myLabel = Label(root, text="DATA ENTRY", bg="#000000", fg="White")
# placing it
myLabel.grid(row=1, column=51)

myLabel1 = Label(root, text="Enter Buyer Name", bg="#000000", fg="White")
# placing it
myLabel1.grid(row=3, column=3)

e = Entry(root)
# for taking input from the user width=50 for changing the width, bg for background fg for letters
e.grid(row=3, column=52, columnspan=3)

myLabel2 = Label(root, text="Enter product", bg="#000000", fg="White")
# placing it
myLabel2.grid(row=4, column=3)

e1 = Entry(root)
# for taking input from the user width=50 for changing the width, bg for background fg for letters
e1.grid(row=4, column=52, columnspan=3)

myLabel3 = Label(root, text="Enter weight", bg="#000000", fg="White")
# placing it
myLabel3.grid(row=6, column=3)

e2 = Entry(root)
# for taking input from the user width=50 for changing the width, bg for background fg for letters
e2.grid(row=6, column=52, columnspan=3)


myLabel4 = Label(root, text="Enter price", bg="#000000", fg="White")
myLabel4.grid(row=8, column=3)

e3 = Entry(root)
# for taking input from the user width=50 for changing the width, bg for background fg for letters
e3.grid(row=8, column=52, columnspan=3)

# creating a button
Mybutton = Button(root, text="Enter", command=clicked, bg="#676767", fg="White")
#Mybutton is for calling a function if a button is clicked
Mybutton.grid(row=9, column=51)


wb.close()

root.mainloop()
